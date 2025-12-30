# coding:utf-8
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QLabel, QFileDialog, QScroller, QScrollerProperties
from qfluentwidgets import FluentIcon as FIF
from qfluentwidgets import qconfig, ScrollArea, PrimaryPushButton, InfoBar, InfoBarPosition, PushButton, MessageBox
from .common.style_sheet import StyleSheet
from .tools.warp_export import warpExport, WarpExport, detect_format, uigf_to_srgf_hkrpg, srgf_to_uigf_hkrpg
import pyperclip
import json
import markdown
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import time


class WarpInterface(ScrollArea):
    def __init__(self, parent=None):
        super().__init__(parent=parent)
        self.view = QWidget(self)
        self.vBoxLayout = QVBoxLayout(self.view)
        self.titleLabel = QLabel("워프 기록", self)

        self.updateBtn = PrimaryPushButton(FIF.SYNC, "데이터 업데이트", self)
        self.updateFullBtn = PushButton(FIF.SYNC, "전체 데이터 업데이트", self)
        self.importBtn = PushButton(FIF.PENCIL_INK, "데이터 가져오기", self)
        self.exportBtn = PushButton(FIF.SAVE_COPY, "데이터 내보내기", self)
        self.exportExcelBtn = PushButton(FIF.SAVE_COPY, "Excel 내보내기", self)
        self.copyLinkBtn = PushButton(FIF.SHARE, "링크 복사", self)
        self.clearBtn = PushButton(FIF.DELETE, "비우기", self)
        self.warplink = None

        self.stateTooltip = None

        self.contentLabel = QLabel(parent)

        qconfig.themeChanged.connect(self.setContent)

        self.setContent()

        self.__initWidget()
        self.__connectSignalToSlot()

    def __initWidget(self):
        self.titleLabel.move(36, 30)
        self.updateBtn.move(35, 80)
        self.updateFullBtn.move(150, 80)
        self.importBtn.move(293, 80)
        self.exportBtn.move(408, 80)
        self.exportExcelBtn.move(523, 80)
        self.copyLinkBtn.move(638, 80)
        self.copyLinkBtn.setEnabled(False)
        self.clearBtn.move(753, 80)

        self.view.setObjectName('view')
        self.setViewportMargins(0, 120, 0, 20)
        self.setObjectName('warpInterface')
        self.contentLabel.setObjectName('contentLabel')
        self.titleLabel.setObjectName('warpLabel')
        StyleSheet.WARP_INTERFACE.apply(self)

        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.setWidget(self.view)
        self.setWidgetResizable(True)
        self.contentLabel.setWordWrap(True)
        self.contentLabel.setMaximumWidth(800)

        # self.vBoxLayout.setSpacing(8)
        # self.vBoxLayout.setAlignment(Qt.AlignTop)
        # self.vBoxLayout.addWidget(self.titleLabel, 0, Qt.AlignTop)

        # self.vBoxLayout.setSpacing(28)
        self.vBoxLayout.setContentsMargins(36, 0, 36, 0)
        self.vBoxLayout.addWidget(self.contentLabel, 0, Qt.AlignTop)

        QScroller.grabGesture(self.viewport(), QScroller.ScrollerGestureType.LeftMouseButtonGesture)
        scroller = QScroller.scroller(self.viewport())
        scroller_props = scroller.scrollerProperties()
        scroller_props.setScrollMetric(QScrollerProperties.ScrollMetric.OvershootDragDistanceFactor, 0.05)
        scroller_props.setScrollMetric(QScrollerProperties.ScrollMetric.OvershootScrollDistanceFactor, 0.05)
        scroller_props.setScrollMetric(QScrollerProperties.ScrollMetric.DecelerationFactor, 0.5)
        scroller.setScrollerProperties(scroller_props)

    def __connectSignalToSlot(self):
        self.updateBtn.clicked.connect(self.__onUpdateBtnClicked)
        self.updateFullBtn.clicked.connect(self.__onUpdateFullBtnClicked)
        self.importBtn.clicked.connect(self.__onImportBtnClicked)
        self.exportBtn.clicked.connect(self.__onExportBtnClicked)
        self.exportExcelBtn.clicked.connect(self.__onExportExcelBtnClicked)
        self.copyLinkBtn.clicked.connect(self.__onCopyLinkBtnClicked)
        self.clearBtn.clicked.connect(self.__onClearBtnClicked)

    def __onUpdateBtnClicked(self):
        warpExport(self)

    def __onUpdateFullBtnClicked(self):
        warpExport(self, "full")

    def __onImportBtnClicked(self):
        try:
            path, _ = QFileDialog.getOpenFileName(self, "워프 기록 가져오기", "", "UIGF / SRGF 형식 (*.json)")
            if not path:
                return

            with open(path, 'r', encoding='utf-8') as file:
                config = json.load(file)
                fmt = detect_format(config)
                if fmt == "uigf":
                    config = uigf_to_srgf_hkrpg(config)
                elif fmt == "neither":
                    raise ValueError("Invalid format")
            warp = WarpExport(config)
            warp.info['export_timestamp'] = int(time.time())
            warp.info['export_app'] = "March7thAssistant"
            try:
                with open("./assets/config/version.txt", 'r', encoding='utf-8') as file:
                    version = file.read()
            except Exception:
                version = ""
            warp.info['export_app_version'] = version
            config = warp.export_data()
            with open("./warp.json", 'w', encoding='utf-8') as file:
                json.dump(config, file, ensure_ascii=False, indent=4)

            self.setContent()

            InfoBar.success(
                title='가져오기 성공(＾∀＾●)',
                content="",
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=1000,
                parent=self
            )
        except Exception:
            InfoBar.warning(
                title='가져오기 실패(╥╯﹏╰╥)',
                content="",
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=1000,
                parent=self
            )

    def __onExportBtnClicked(self):
        try:
            with open("./warp.json", 'r', encoding='utf-8') as file:
                config = json.load(file)
            warp = WarpExport(config)
            default_name = f"UIGF_{warp.get_uid()}"

            path, _ = QFileDialog.getSaveFileName(
                self,
                "워프 기록 내보내기",
                f"{default_name}",
                "UIGF 형식 (*.json);;SRGF 형식 (*.srgf.json)"
            )
            if not path:
                return

            is_srgf = path.lower().endswith(".srgf.json")

            if is_srgf:
                data_to_save = config
            else:
                # SRGF → UIGF
                data_to_save = srgf_to_uigf_hkrpg(config)

            with open(path, 'w', encoding='utf-8') as file:
                json.dump(data_to_save, file, ensure_ascii=False, indent=4)

            os.startfile(os.path.dirname(path))

            InfoBar.success(
                title='내보내기 성공(＾∀＾●)',
                content="",
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=1000,
                parent=self
            )

        except Exception:
            InfoBar.warning(
                title='내보내기 실패(╥╯﹏╰╥)',
                content="",
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=1000,
                parent=self
            )

    def __onExportExcelBtnClicked(self):
        try:
            with open("./warp.json", 'r', encoding='utf-8') as file:
                config = json.load(file)
            records = config.get("list", [])
            df = pd.DataFrame(records)
            df = df[["time", "name", "item_type", "rank_type", "gacha_type"]]
            gacha_map = {
                "11": "캐릭터 이벤트 워프",
                "12": "광추 이벤트 워프",
                "21": "캐릭터 컬래버 워프",
                "22": "광추 컬래버 워프",
                "1": "일반 워프",
                "2": "초행길 워프",
            }
            df["gacha_type"] = df["gacha_type"].map(gacha_map).fillna("알 수 없음")
            df.rename(columns={
                "time": "시간",
                "name": "이름",
                "item_type": "유형",
                "rank_type": "등급",
                "gacha_type": "워프 유형",
            }, inplace=True)
            df["총 횟수"] = range(1, len(df) + 1)
            df["천장 내 횟수"] = 0
            pity_counters = {}
            for idx, row in df.iterrows():
                pool = row["워프 유형"]
                star = row["등급"]
                if pool not in pity_counters:
                    pity_counters[pool] = 0
                pity_counters[pool] += 1
                df.at[idx, "천장 내 횟수"] = pity_counters[pool]
                if star == "5":
                    pity_counters[pool] = 0
            path, _ = QFileDialog.getSaveFileName(
                self,
                "Excel 파일로 내보내기",
                f"워프_기록_{config['info'].get('uid', '알수없음')}.xlsx",
                "Excel 파일 (*.xlsx)"
            )
            if not path:
                return

            df.to_excel(path, index=False)
            wb = load_workbook(path)
            ws = wb.active
            for row in range(2, ws.max_row + 1):
                star_cell = ws[f"D{row}"]
                try:
                    star = star_cell.value
                    if star == "5":
                        for col in ws[row]:
                            col.font = Font(color="FFA500")
                    elif star == "4":
                        for col in ws[row]:
                            col.font = Font(color="800080")
                except:
                    continue

            for column_cells in ws.columns:
                max_width = 0
                col_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    try:
                        value = str(cell.value) if cell.value else ""
                        width = 0
                        for ch in value:
                            if u'\u4e00' <= ch <= u'\u9fff':
                                width += 2
                            else:
                                width += 1
                        if width > max_width:
                            max_width = width
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_width + 2

            wb.save(path)
            os.startfile(os.path.dirname(path))

            InfoBar.success(
                title='내보내기 성공(＾∀＾●)',
                content="",
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=1000,
                parent=self
            )

        except Exception:
            InfoBar.warning(
                title='내보내기 실패(╥╯﹏╰╥)',
                content="",
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=1000,
                parent=self
            )

    def __onCopyLinkBtnClicked(self):
        try:
            pyperclip.copy(self.warplink)
            InfoBar.success(
                title='복사 성공(＾∀＾●)',
                content="",
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=1000,
                parent=self
            )
        except Exception:
            InfoBar.warning(
                title='복사 실패(╥╯﹏╰╥)',
                content="",
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=1000,
                parent=self
            )

    def __onClearBtnClicked(self):
        message_box = MessageBox(
            "워프 기록 비우기",
            "정말 워프 기록을 비우시겠습니까? 이 작업은 취소할 수 없습니다.",
            self.window()
        )
        message_box.yesButton.setText('확인')
        message_box.cancelButton.setText('취소')
        if message_box.exec():
            try:
                os.remove("./warp.json")
                self.setContent()
                InfoBar.success(
                    title='비우기 완료(＾∀＾●)',
                    content="",
                    orient=Qt.Horizontal,
                    isClosable=True,
                    position=InfoBarPosition.TOP,
                    duration=1000,
                    parent=self
                )
            except Exception as e:
                print(e)
                InfoBar.warning(
                    title='비우기 실패(╥╯﹏╰╥)',
                    content="",
                    orient=Qt.Horizontal,
                    isClosable=True,
                    position=InfoBarPosition.TOP,
                    duration=1000,
                    parent=self
                )

    def setContent(self):
        try:
            with open("./warp.json", 'r', encoding='utf-8') as file:
                config = json.load(file)

            warp = WarpExport(config)
            if qconfig.theme.name == "DARK":
                content = warp.data_to_html("dark")
            else:
                content = warp.data_to_html("light")
            self.clearBtn.setEnabled(True)
            self.exportBtn.setEnabled(True)
            self.exportExcelBtn.setEnabled(True)
        except Exception as e:
            content = "워프 기록이 비어있습니다. 먼저 게임 내 워프 기록을 연 다음, 데이터 업데이트를 클릭하세요.\n\nStarRail Warp Export 또는 Starward와 같은 UIGF/SRGF 데이터 형식을 지원하는 다른 앱에서 데이터를 가져올 수도 있습니다.\n\n링크 복사 기능은 미니 프로그램이나 다른 소프트웨어에서 사용할 수 있습니다."
            self.clearBtn.setEnabled(False)
            self.exportBtn.setEnabled(False)
            self.exportExcelBtn.setEnabled(False)
        self.contentLabel.setText(content)